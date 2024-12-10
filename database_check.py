import csv
import random
import numpy as np
import openpyxl
from openpyxl import Workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import pandas as pd

st = ['SMILES', 'Microbe', 'Y']

aBiofilm_drug_smiles = []
workbook = openpyxl.load_workbook('./source_dealed_data/aBiofilm_data/drug_microbe_matrix.xlsx')
sheet = workbook.active
dataa = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    aBiofilm_drug_smiles.append(row_data[1])
    dataa.append(row_data)
aBiofilm_microbe_name = dataa[0]

del aBiofilm_microbe_name[0]

MDAD_drug_smiles = []
workbook = openpyxl.load_workbook('./source_dealed_data/MDAD_data/drug_microbe_matrix.xlsx')
sheet = workbook.active
data = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    MDAD_drug_smiles.append(row_data[1])
    data.append(row_data)
MDAD_microbe_name = data[0]

del MDAD_microbe_name[0]

MDAD_index = np.loadtxt("./data_index/MDAD_index/index.txt").tolist()
aBiofilm_index = np.loadtxt("./data_index/aBiofilm_index/index.txt").tolist()

MDAD_repeated_drugs = []
MDAD_only_drugs = []
for i in range(len(MDAD_index)):
    if MDAD_drug_smiles[int(MDAD_index[i])] in aBiofilm_drug_smiles:
        MDAD_repeated_drugs.append(int(MDAD_index[i]))
    else:
        MDAD_only_drugs.append(int(MDAD_index[i]))

MDAD_repeated_microbes = []
MDAD_only_microbes = []
for i in range(1, 174):
    if MDAD_microbe_name[i] in aBiofilm_microbe_name:
        MDAD_repeated_microbes.append(i)
    else:
        MDAD_only_microbes.append(i)

aBiofilm_repeated_microbes = []
aBiofilm_only_microbes = []
for i in range(1, 141):
    if aBiofilm_microbe_name[i] in MDAD_microbe_name:
        aBiofilm_repeated_microbes.append(i)
    else:
        aBiofilm_only_microbes.append(i)

aBiofilm_selected_drug_index = []
selected_d = [88, 90, 91, 107, 109, 112, 113, 128, 149, 158, 160, 165, 202, 211, 221, 222, 224, 225, 229, 230, 231] #selected MDAD drug index
print(MDAD_drug_smiles)
print(aBiofilm_drug_smiles)
for i in selected_d:
    aBiofilm_selected_drug_index.append(aBiofilm_drug_smiles.index(MDAD_drug_smiles[i]))

print("MDAD_repeated_drugs:" + str(len(MDAD_repeated_drugs)) + str(MDAD_repeated_drugs))
print("MDAD_only_drugs:" + str(len(MDAD_only_drugs)) + str(MDAD_only_drugs))
print("MDAD_repeated_microbes:" + str(len(MDAD_repeated_microbes)) + str(MDAD_repeated_microbes))
print("MDAD_only_microbes:" + str(len(MDAD_only_microbes)) + str(MDAD_only_microbes))
print("aBiofilm_only_microbes:" + str(len(aBiofilm_only_microbes)) + str(aBiofilm_only_microbes))
print("aBiofilm_selected_drug_index:" + str(len(aBiofilm_selected_drug_index)) + str(aBiofilm_selected_drug_index))

workbook = openpyxl.load_workbook('./source_dealed_data/MDAD_data/microbes.xlsx')
sheet = workbook.active
mi_data = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    mi_data.append(row_data)
workbook = openpyxl.load_workbook('./source_dealed_data/aBiofilm_data/microbes.xlsx')
sheet = workbook.active
mi_data_2 = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    mi_data_2.append(row_data)

mi_r_data = []
for mi in MDAD_repeated_microbes:
    mi_r_data.append([mi_data[mi][0], mi_data[mi][1], mi_data[mi][2], mi_data[mi][4], mi_data[mi][6]])

workbook = Workbook()
worksheet = workbook.active
for i in range(len(mi_data[0])):
    worksheet.cell(row=1, column=1 + i).value = mi_data[0][i]
for p in range(len(mi_r_data)):
    for q in range(len(mi_r_data[0])):
        worksheet.cell(row=2 + p, column=1 + q).value = mi_r_data[p][q]
workbook.save("./source_dealed_data/repeated_microbes.xlsx")

mn = []
mi_a_data = []
for mi in range(174):
    mi_a_data.append([mi_data[mi][0], mi_data[mi][1], mi_data[mi][2], mi_data[mi][4], mi_data[mi][6]])
    mn.append(mi_data[mi][1])

for mi in aBiofilm_only_microbes:
    mi_a_data.append([mi_data_2[mi][0], mi_data_2[mi][1], mi_data_2[mi][2], mi_data_2[mi][4], mi_data_2[mi][6]])
    mn.append(mi_data_2[mi][1])
del mn[0]

workbook = Workbook()
worksheet = workbook.active
for p in range(len(mi_a_data)):
    for q in range(5):
        worksheet.cell(row=1 + p, column=1 + q).value = mi_a_data[p][q]
workbook.save("./source_dealed_data/all_microbes.xlsx")

data = pd.read_excel('./source_dealed_data/all_microbes.xlsx')
filtered_data = data
tfidf_vectorizer = TfidfVectorizer()
tfidf_matrix = tfidf_vectorizer.fit_transform(filtered_data['Microbe_name'])
cosine_sim_matrix = cosine_similarity(tfidf_matrix)
np.savetxt("./source_dealed_data/all_microbe_cos.txt", cosine_sim_matrix)
with open("./source_dealed_data/all_microbe_name.txt", 'w', encoding='utf-8') as f:
    for i in range(len(mn)):
        f.write(mn[i]+'\n')

#venne
workbook = openpyxl.load_workbook('./source_dealed_data/aBiofilm_data/drugs.xlsx')
sheet = workbook.active
aBiofilm_drug_name = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    aBiofilm_drug_name.append(row_data[1])
workbook = openpyxl.load_workbook('./source_dealed_data/MDAD_data/drugs.xlsx')
sheet = workbook.active
MDAD_drug_name = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    MDAD_drug_name.append(row_data[1])

MDAD_known = np.loadtxt("./data_index/MDAD_index/known.txt").tolist()
MDAD_known = [[MDAD_drug_name[int(row[0])], MDAD_microbe_name[int(row[1])]] for row in MDAD_known]
aBiofilm_known = np.loadtxt("./data_index/aBiofilm_index/known.txt").tolist()
aBiofilm_known = [[aBiofilm_drug_name[int(row[0])], aBiofilm_microbe_name[int(row[1])]] for row in aBiofilm_known]
label_1_set = MDAD_known
for value in aBiofilm_known:
    if value not in MDAD_known:
        label_1_set.append(value)

workbook = openpyxl.load_workbook('./new_probably_connections(Fuzzy_set)/aBiofilm/connections.xlsx')
sheet = workbook.active
aBiofilm_pairs = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    aBiofilm_pairs.append([row_data[0], row_data[1]])
workbook = openpyxl.load_workbook('./new_probably_connections(Fuzzy_set)/MDAD/connections.xlsx')
sheet = workbook.active
MDAD_pairs = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    MDAD_pairs.append([row_data[0], row_data[1]])

fuzzy_set = MDAD_pairs
for value in aBiofilm_pairs:
    if value not in MDAD_pairs:
        fuzzy_set.append(value)

print('Associations amount: '+ str(len(label_1_set)))
print('Fuzzy set amount: '+ str(len(fuzzy_set)))
a = 0
for value in fuzzy_set:
    if value in label_1_set:
        a += 1
print('Repeated amount for associations and fuzzy set: '+ str(a))