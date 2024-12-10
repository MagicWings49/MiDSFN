import requests
from Bio import Entrez
import openpyxl
import numpy as np
from openpyxl import Workbook

# Set entrez email
Entrez.email = "example@qq.com"

workbook = openpyxl.load_workbook("microbes.xlsx")
sheet = workbook.active
MD_m = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    MD_m.append(row_data[1])
del MD_m[0]
MD_m = MD_m[:173]

workbook = openpyxl.load_workbook("drugs.xlsx")
sheet = workbook.active
MD_d = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    MD_d.append(row_data[1])
del MD_d[0]
MD_d = MD_d[:1373]

D = np.loadtxt("./index.txt")
print(D)
D = D[:]   # If you find the code cannot run all the searching in one time,
           # then start from the place where you end last time and change the saving path in line 83,
           # finally montage all parts together.               Runtime:2024/9/23 to 2024/10/1

def search_pubmed(microbe, drug):
    query = f"{microbe} AND {drug}"
    handle = Entrez.esearch(db="pubmed", term=query, retmax=10)
    record = Entrez.read(handle)
    handle.close()
    return record["IdList"]

def fetch_pubmed_details(pubmed_ids):
    ids = ",".join(pubmed_ids)
    handle = Entrez.efetch(db="pubmed", id=ids, rettype="medline", retmode="text")
    records = handle.read()
    handle.close()
    return records

known = []

def main(a, b, c, d):
    drug = a
    microbe = b
    pubmed_ids = search_pubmed(microbe, drug)
    if pubmed_ids:
        #details = fetch_pubmed_details(pubmed_ids)
        print(f"PubMed IDs: {pubmed_ids}" + "////" + str(d) + "////" + b)
        #print(f"Details: {details}")
        known.append([a, b, str(pubmed_ids), c, d])
    #else:
        #print("No related PubMed articles found.")

k = 1
for i in range(len(D)):
    print(MD_d[int(D[i]-1)]+"------------------------------"+"No."+str(int(D[i])))
    for j in range(len(MD_m)):
        main(MD_d[int(D[i]-1)], MD_m[j], int(D[i]), j+1)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=1, column=1).value = "Drug"
    worksheet.cell(row=1, column=2).value = "Microbe"
    worksheet.cell(row=1, column=3).value = "PubMed_ID"
    worksheet.cell(row=1, column=4).value = "Drug_index"
    worksheet.cell(row=1, column=5).value = "Microbe_index"

    for p in range(len(known)):
        for q in range(len(known[0])):
            worksheet.cell(row=2 + p, column=1 + q).value = known[p][q]

    workbook.save("./connections14.xlsx")
    k += 1
