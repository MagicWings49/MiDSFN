import requests
from Bio import Entrez
import openpyxl
import numpy as np
from openpyxl import Workbook

# Set entrez email
Entrez.email = "example@qq.com"

workbook = openpyxl.load_workbook("microbes.xlsx")
sheet = workbook.active
aB_m = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    aB_m.append(row_data[1])
del aB_m[0]
aB_m = aB_m[:140]

workbook = openpyxl.load_workbook("drugs.xlsx")
sheet = workbook.active
aB_d = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    aB_d.append(row_data[1])
del aB_d[0]
aB_d = aB_d[:1720]

D = np.loadtxt("./index.txt")
print(D)
D = D[:]   # If you find the code cannot run all the searching in one time,
           # then start from the place where you end last time and change the saving path in line 83,
           # finally montage all parts together.               Runtime:2024/9/23 to 2024/10/1

def search_pubmed(microbe, drug):
    query = f"{microbe} AND {drug}"
    handle = Entrez.esearch(db="pubmed", term=query, retmax=10)
    record = Entrez.read(handle)
    handle.close()
    return record["IdList"]

def fetch_pubmed_details(pubmed_ids):
    ids = ",".join(pubmed_ids)
    handle = Entrez.efetch(db="pubmed", id=ids, rettype="medline", retmode="text")
    records = handle.read()
    handle.close()
    return records

known = []

def main(a, b, c, d):
    drug = a
    microbe = b
    pubmed_ids = search_pubmed(microbe, drug)
    if pubmed_ids:
        details = fetch_pubmed_details(pubmed_ids)
        print(f"PubMed IDs: {pubmed_ids}" + "////" + str(d) + "////" + b)
        #print(f"Details: {details}")
        known.append([a, b, str(pubmed_ids), c, d])
    #else:
        #print("No related PubMed articles found.")

k = 1
for i in range(len(D)):
    print(aB_d[int(D[i]-1)]+"------------------------------"+"No."+str(int(D[i])))
    for j in range(len(aB_m)):
        main(aB_d[int(D[i]-1)], aB_m[j], int(D[i]), j+1)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=1, column=1).value = "Drug"
    worksheet.cell(row=1, column=2).value = "Microbe"
    worksheet.cell(row=1, column=3).value = "PubMed_ID"
    worksheet.cell(row=1, column=4).value = "Drug_index"
    worksheet.cell(row=1, column=5).value = "Microbe_index"

    for p in range(len(known)):
        for q in range(len(known[0])):
            worksheet.cell(row=2 + p, column=1 + q).value = known[p][q]

    workbook.save("./connections1.xlsx")
    k += 1
