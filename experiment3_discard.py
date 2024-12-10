import csv
import openpyxl

workbook = openpyxl.load_workbook('new_probably_connections(Fuzzy_set)/MDAD/connections.xlsx')
sheet = workbook.active
MD_probably_index = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    MD_probably_index.append([row_data[3], row_data[4]])
del MD_probably_index[0]

workbook = openpyxl.load_workbook('new_probably_connections(Fuzzy_set)/aBiofilm/connections.xlsx')
sheet = workbook.active
aB_probably_index = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    aB_probably_index.append([row_data[3], row_data[4]])
del aB_probably_index[0]

MDAD_drug_smiles = []
workbook = openpyxl.load_workbook('./source_dealed_data/MDAD_data/drug_microbe_matrix.xlsx')
sheet = workbook.active
dataa = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    MDAD_drug_smiles.append(row_data[1])
    dataa.append(row_data)
MDAD_microbe_name = dataa[0]

aBiofilm_drug_smiles = []
workbook = openpyxl.load_workbook('./source_dealed_data/aBiofilm_data/drug_microbe_matrix.xlsx')
sheet = workbook.active
data = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    aBiofilm_drug_smiles.append(row_data[1])
    data.append(row_data)
aBiofilm_microbe_name = data[0]

st = ['SMILES', 'Microbe', 'Y']

for set_ in ['train', 'val', 'test']:
    with open('./datasets/drug_microbe/MDAD_aBiofilm/' + set_ + '.csv', 'r', newline='\n', encoding='utf-8') as file:
        reader = csv.reader(file)
        dataset = [row for row in reader]

    file.close()
    dataset_dicard = []
    for i in range(1, len(dataset)):
        if dataset[i][2] == '0':
            if dataset[i][0] in MDAD_drug_smiles:
                if dataset[i][1] in MDAD_microbe_name:
                    if [MDAD_drug_smiles.index(dataset[i][0]), MDAD_microbe_name.index(dataset[i][1])-1] not in MD_probably_index:
                        dataset_dicard.append(dataset[i])
                else:
                    if [aBiofilm_drug_smiles.index(dataset[i][0]), aBiofilm_microbe_name.index(dataset[i][1]) - 1] not in aB_probably_index:
                        dataset_dicard.append(dataset[i])
            else:
                if dataset[i][1] in aBiofilm_microbe_name:
                    if [aBiofilm_drug_smiles.index(dataset[i][0]), aBiofilm_microbe_name.index(dataset[i][1]) - 1] not in aB_probably_index:
                        dataset_dicard.append(dataset[i])
        else:
            dataset_dicard.append(dataset[i])
    with open('./datasets/drug_microbe/MDAD_aBiofilm_discard/' + set_ + '.csv', 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(st)
        for row in dataset_dicard:
            writer.writerow(row)
    f.close()


