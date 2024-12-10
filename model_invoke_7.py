from models import DrugBAN
from utils import set_seed, graph_collate_func, mkdir
from configs import get_cfg_defaults
from dataloader import DTIDataset, MultiDataLoader
from torch.utils.data import DataLoader
import torch
import torch.nn as nn
import argparse
import warnings, os
import torch.nn.functional as F
import openpyxl
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import AgglomerativeClustering, DBSCAN
import matplotlib.pyplot as plt
from sklearn.metrics.pairwise import cosine_similarity, euclidean_distances
from scipy.cluster.hierarchy import dendrogram, linkage
from sklearn.manifold import TSNE
import numpy as np
from scipy.stats import wilcoxon
from openpyxl import Workbook
from scipy.spatial import distance
from scipy.cluster.hierarchy import linkage
from scipy.linalg import LinAlgError, pinv

#Parameters setting: --cfg "configs/MiDSFN.yaml" --data "drug_microbe" --split "Case_prediction"
#Decoder setting in './configs/MiDSFN.yaml': "KAN"

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

parser = argparse.ArgumentParser(description="MiDSFN for Drug-Microbe prediction")
parser.add_argument('--cfg', required=True, type=str, help="path to config file", metavar='CFG')
parser.add_argument('--data', required=True, type=str, metavar='TASK', help='dataset')
parser.add_argument('--split', default='random', type=str, metavar='S', help="split task",
                    choices=['MDAD', 'aBiofilm', 'MDAD_aBiofilm', 'MDAD_aBiofilm_discard', 'aBiofilm_MDAD',
                             'MDAD+aBiofilm', 'MDAD+aBiofilm_discard', 'aBiofilm_to_MDAD', 'Case_prediction',
                             'Ethnic_interaction'])
args = parser.parse_args()

def binary_cross_entropy(pred_output, labels):
    loss_fct = torch.nn.BCELoss()
    m = nn.Sigmoid()
    n = torch.squeeze(m(pred_output), 1)
    loss = loss_fct(n, labels)
    return n, loss

def cross_entropy_logits(linear_output, label, weights=None):
    class_output = F.log_softmax(linear_output, dim=1)
    n = F.softmax(linear_output, dim=1)[:, 1]
    max_class = class_output.max(1)
    y_hat = max_class[1]  # get the index of the max log-probability
    if weights is None:
        loss = nn.NLLLoss()(class_output, label.type_as(y_hat).view(label.size(0)))
    else:
        losses = nn.NLLLoss(reduction="none")(class_output, label.type_as(y_hat).view(label.size(0)))
        loss = torch.sum(weights * losses) / torch.sum(weights)
    return n, loss

cfg = get_cfg_defaults()
cfg.merge_from_file(args.cfg)
ex_split = str(args.split)
n_class = cfg["DECODER"]["BINARY"]
decoder = cfg["DECODER"]["NAME"]
params = {'batch_size': 1, 'shuffle': False, 'num_workers': cfg.SOLVER.NUM_WORKERS,
              'drop_last': True, 'collate_fn': graph_collate_func}
dataFolder = f'./datasets/{args.data}/'
dataFolder = os.path.join(dataFolder, ex_split)
result_path = os.path.join('./result/' + ex_split, decoder)
test_path = os.path.join(dataFolder, "test.csv")
df_test = pd.read_csv(test_path)
test_dataset = DTIDataset(df_test.index.values, df_test)
test_generator = DataLoader(test_dataset, **params)

test_path_2 = os.path.join(dataFolder, "test_2.csv")
df_test_2 = pd.read_csv(test_path_2)
test_dataset_2 = DTIDataset(df_test_2.index.values, df_test_2)
test_generator_2 = DataLoader(test_dataset_2, **params)

MyModel = DrugBAN(**cfg).to(device)
MyModel.eval()
MyModel.load_state_dict(torch.load(result_path + '/best_model_epoch_20.pth')) #This place need to be confirm every time.


def find_top_n_rankings(numbers, n):
    sorted_indices = sorted(range(1, len(numbers)+1), key=lambda i: numbers[i-1], reverse=True)
    rankings = {rank: index for rank, index in enumerate(sorted_indices[:n], start=1)}
    return rankings

def invoke(MyModel, test_generator):
    test_loss = 0
    y_label, y_pred = [], []
    with torch.no_grad():
        t = 0
        for i, (v_d, v_p, a_m, d_i, labels) in enumerate(test_generator):
            a_m = torch.stack(a_m, 0)
            v_d, v_p, a_m, d_i, labels = v_d.to(device), v_p.to(device), a_m.to(device), d_i, labels.float().to(device)
            v_d, v_p, f, score = MyModel(v_d, v_p, a_m, d_i)
            if n_class == 1:
                n, loss = binary_cross_entropy(score, labels)
            else:
                n, loss = cross_entropy_logits(score, labels)
            test_loss += loss.item()
            y_label = y_label + labels.to("cpu").tolist()
            y_pred = y_pred + n.to("cpu").tolist()
            if labels == 1:
                t += 1
    return t, y_pred, y_label

n_con, scores, _ = invoke(MyModel, test_generator)

print(scores[:n_con])
rankings = find_top_n_rankings(scores, n_con)

print(rankings)

n_con_2, scores_2, labels_2 = invoke(MyModel, test_generator_2)
microbe_scores = [[scores_2[j*62 + i] for j in range(21)] for i in range(62)]
microbe_labels = [[labels_2[j*62 + i] for j in range(21)] for i in range(62)]

selected_m = [1, 2, 3, 4, 5, 6, 8, 10, 11, 12, 16, 18, 20, 22, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39,
              40, 41, 42, 43, 44, 49, 50, 51, 52, 56, 58, 60, 61, 63, 65, 66, 67, 68, 72, 73, 74, 76, 77, 78, 79,
              94, 95, 96, 97, 98, 99, 100, 103, 104, 105, 106, 108]

def clusters(type, dis):
    data = pd.read_excel('./source_dealed_data/repeated_microbes.xlsx')
    filtered_data = data[data['Microbe_ID'] <= 108]
    if type == 'Txt':
        tfidf_vectorizer = TfidfVectorizer()
        tfidf_matrix = tfidf_vectorizer.fit_transform(filtered_data['Microbe_name'])
        euclidean_dis_matrix = euclidean_distances(tfidf_matrix.toarray())
        mahalanobis_dis_matrix = distance.cdist(tfidf_matrix.toarray(), tfidf_matrix.toarray(), 'mahalanobis',
                                                VI=np.linalg.inv(np.cov(tfidf_matrix.toarray(), rowvar=False)))
        ii = 'a'
    elif type == 'Genomic_embedding':
        microbe_gene_emb = np.loadtxt('./result/Case_prediction/microbal_genomic_embedding.txt').tolist()
        microbe_gene_emb = [microbe_gene_emb[s - 1] for s in selected_m]
        euclidean_dis_matrix = euclidean_distances(microbe_gene_emb)
        mahalanobis_dis_matrix = distance.cdist(microbe_gene_emb, microbe_gene_emb, 'mahalanobis',
                                                VI=np.linalg.inv(np.cov(microbe_gene_emb, rowvar=False)))
        ii = 'b'
    elif type == 'Scores':
        euclidean_dis_matrix = euclidean_distances(np.array(microbe_scores))
        mahalanobis_dis_matrix = distance.cdist(np.array(microbe_scores), np.array(microbe_scores), 'mahalanobis',
                                                VI=np.linalg.inv(np.cov(np.array(microbe_scores), rowvar=False)))
        ii = 'c'
    elif type == 'Labels':
        euclidean_dis_matrix = euclidean_distances(np.array(microbe_labels))
        try:
            cov_mat = np.cov(microbe_labels, rowvar=False)
            inv_cov_mat = np.linalg.inv(cov_mat)
        except LinAlgError:
            inv_cov_mat = pinv(cov_mat)
        mahalanobis_dis_matrix = distance.cdist(np.array(microbe_labels), np.array(microbe_labels), 'mahalanobis',
                                                VI=inv_cov_mat)
        ii = 'd'
    if dis == 'mah':
        dis_matrix = mahalanobis_dis_matrix
    elif dis == 'euc':
        dis_matrix = euclidean_dis_matrix

    #np.savetxt('./dis_matrix_' + str(type) + '.txt', dis_matrix)
    clustering = AgglomerativeClustering(n_clusters=None, distance_threshold=0.7)
    clustering.fit(dis_matrix)
    labels = clustering.labels_
    Z = linkage(dis_matrix, 'ward')
    plt.figure(figsize=(18, 9))
    plt.title(ii + '. Hierarchical Clustering Heatmap for Microbal ' + str(type))
    dendrogram(Z, orientation='right', labels=filtered_data['Microbe_name'].values, show_leaf_counts=True)
    if type == 'Labels':
        plt.xlim(-2, 170)
    plt.savefig('./Cluster_fig/Euc(Mah)+Ward/' + str(type) + '.png', dpi=400)
    plt.show()


def dbscan(type):
    data = pd.read_excel('./source_dealed_data/repeated_microbes.xlsx')
    tsne = TSNE(n_components=2, perplexity=10, learning_rate=200, random_state=0)
    filtered_data = data[data['Microbe_ID'] <= 108]
    microbe_data = filtered_data['Microbe_name'].tolist()
    if type == 'Txt':
        tfidf_vectorizer = TfidfVectorizer()
        tfidf_matrix = tfidf_vectorizer.fit_transform(filtered_data['Microbe_name'])
        eps = 26          #33
        ii = 'a'
    elif type == 'Genomic_embedding':
        microbe_gene_emb = np.loadtxt('./result/Case_prediction/microbal_genomic_embedding.txt').tolist()
        tfidf_matrix = [microbe_gene_emb[s - 1] for s in selected_m]
        eps = 24
        ii = 'b'
    elif type == 'Scores':
        tfidf_matrix = microbe_scores
        eps = 30
        ii = 'c'
    elif type == 'Labels':
        tfidf_matrix = microbe_labels
        eps = 50
        ii = 'c'
    tfidf_matrix = tsne.fit_transform(tfidf_matrix)
    db = DBSCAN(eps=eps, min_samples=4).fit(tfidf_matrix)
    labels = db.labels_
    unique_labels = set(labels)
    plt.figure(figsize=(18, 9))
    colors = [plt.cm.Spectral(each) for each in np.linspace(0, 1, len(unique_labels))]
    if type == 'Genomic_embedding':
        colors[3] = (0, 0, 1, 1)
    elif type == 'Scores':
        colors[2] = (0, 0, 1, 1)
    print(colors)
    for k, col in zip(unique_labels, colors):
        if k == -1:
            col = [0, 0, 0, 1]
        class_member_mask = (labels == k)
        xy = tfidf_matrix[class_member_mask, :]
        microbe_name = [microbe_data[s] for s in [i for i, x in enumerate(class_member_mask) if x]]
        plt.scatter(xy[:, 0], xy[:, 1], c=col, edgecolor=col, linewidths=0.35, s=9)
        for j in range(len(xy)):
            plt.annotate(microbe_name[j], (xy[j, 0], xy[j, 1]), c=col, textcoords="offset points", xytext=(0, 2),
                         ha='center', fontsize=6, zorder=0.5)
    plt.title(ii + '. DBSCAN Clustering for Microbal ' + str(type))
    if type == 'Labels':
        plt.xlim(-318, 516)
        plt.ylim(-365, 310)

    plt.savefig('./Cluster_fig/t-SNE+DBSCAN/' + str(type) + '.png', dpi=400)
    plt.show()


clusters('Txt', 'euc')
clusters('Genomic_embedding', 'euc')
clusters('Scores', 'mah')
clusters('Labels', 'mah')

dbscan('Txt')
dbscan('Genomic_embedding')
dbscan('Scores')
dbscan('Labels')

workbook = openpyxl.load_workbook('./source_dealed_data/repeated_microbes.xlsx')
sheet = workbook.active
microbes = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    microbes.append(row_data[1])
microbes = microbes[0:63]
workbook = openpyxl.load_workbook('./new_probably_connections(Fuzzy_set)/MDAD/connections0.xlsx')
sheet = workbook.active
drugs = ['Drug']
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    if row_data[0] != drugs[-1]:
        drugs.append(row_data[0])

def pairwise_wilcoxon_test(input):
    m, n = input.shape
    p_matrix = np.zeros((m, m))
    for i in range(m):
        for j in range(m):
            if i != j:
                sample1 = input[i]
                sample2 = input[j]
                w_statistic, p_value = wilcoxon(sample1, sample2)
                p_matrix[i, j] = p_value
            else:
                p_matrix[i, j] = 1
    return p_matrix.tolist()

print('Wilcoxon Signed-Rank Test:')
p_split = [1, 0.05, 0.01, 0.001, 0.0001]
D_p = pairwise_wilcoxon_test(np.array(microbe_scores).T)
count1 = sum(1 for row in D_p for value in row if p_split[1] <= float(value) < p_split[0])
count2 = sum(1 for row in D_p for value in row if p_split[2] <= float(value) < p_split[1])
count3 = sum(1 for row in D_p for value in row if p_split[3] <= float(value) < p_split[2])
count4 = sum(1 for row in D_p for value in row if p_split[4] <= float(value) < p_split[3])
count5 = sum(1 for row in D_p for value in row if float(value) < p_split[4])
print('Proportion for drugs_p between ' + str(p_split[1]) + ' and ' + str(p_split[0]) + ': ' + str(count1/(21*20)))
print('Proportion for drugs_p between ' + str(p_split[2]) + ' and ' + str(p_split[1]) + ': ' + str(count2/(21*20)))
print('Proportion for drugs_p between ' + str(p_split[3]) + ' and ' + str(p_split[2]) + ': ' + str(count3/(21*20)))
print('Proportion for drugs_p between ' + str(p_split[4]) + ' and ' + str(p_split[3]) + ': ' + str(count4/(21*20)))
print('Proportion for drugs_p between ' + str(0) + ' and ' + str(p_split[4]) + ': ' + str(count5/(21*20)))
D_p.insert(0, drugs)
for i in range(1, len(D_p)):
    D_p[i].insert(0, drugs[i])

M_p = pairwise_wilcoxon_test(np.array(microbe_scores))
count1 = sum(1 for row in M_p for value in row if p_split[1] <= float(value) < p_split[0])
count2 = sum(1 for row in M_p for value in row if p_split[2] <= float(value) < p_split[1])
count3 = sum(1 for row in M_p for value in row if p_split[3] <= float(value) < p_split[2])
count4 = sum(1 for row in M_p for value in row if p_split[4] <= float(value) < p_split[3])
count5 = sum(1 for row in M_p for value in row if float(value) < p_split[4])
print('Proportion for microbe_p between ' + str(p_split[1]) + ' and ' + str(p_split[0]) + ': ' + str(count1/(62*61)))
print('Proportion for microbe_p between ' + str(p_split[2]) + ' and ' + str(p_split[1]) + ': ' + str(count2/(62*61)))
print('Proportion for microbe_p between ' + str(p_split[3]) + ' and ' + str(p_split[2]) + ': ' + str(count3/(62*61)))
print('Proportion for microbe_p between ' + str(p_split[4]) + ' and ' + str(p_split[3]) + ': ' + str(count4/(62*61)))
print('Proportion for microbe_p between ' + str(0) + ' and ' + str(p_split[4]) + ': ' + str(count5/(62*61)))
M_p.insert(0, microbes)
for i in range(1, len(M_p)):
    M_p[i].insert(0, microbes[i])

workbook = Workbook()
worksheet = workbook.active
for p in range(len(D_p)):
    for q in range(len(D_p[0])):
        worksheet.cell(row=1 + p, column=1 + q).value = D_p[p][q]
workbook.save("./result/Case_prediction/Drug_wilcoxon_p.xlsx")

workbook = Workbook()
worksheet = workbook.active
for p in range(len(M_p)):
    for q in range(len(M_p[0])):
        worksheet.cell(row=1 + p, column=1 + q).value = M_p[p][q]
workbook.save("./result/Case_prediction/Microbe_wilcoxon_p.xlsx")

