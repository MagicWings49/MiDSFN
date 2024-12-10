from models import DrugBAN
from utils import set_seed, graph_collate_func, mkdir
from configs import get_cfg_defaults
from dataloader import DTIDataset, MultiDataLoader
from torch.utils.data import DataLoader
import torch
import torch.nn as nn
import argparse
import warnings, os
import pandas as pd
import torch.nn.functional as F
from rdkit import Chem
from rdkit.Chem import Draw
from PIL import Image, ImageDraw, ImageFont
from rdkit.Chem import AllChem
from openpyxl import Workbook
import numpy as np

#Parameters setting: --cfg "configs/MiDSFN.yaml" --data "drug_microbe" --split "MDAD+aBiofilm_discard"
#Decoder setting in './configs/MiDSFN.yaml': "KAN" or "MLP"

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

parser = argparse.ArgumentParser(description="MiDSFN for Drug-Microbe prediction")
parser.add_argument('--cfg', required=True, type=str, help="path to config file", metavar='CFG')
parser.add_argument('--data', required=True, type=str, metavar='TASK', help='dataset')
parser.add_argument('--split', default='random', type=str, metavar='S', help="split task",
                    choices=['MDAD', 'aBiofilm', 'MDAD_aBiofilm', 'MDAD_aBiofilm_discard', 'aBiofilm_MDAD',
                             'MDAD+aBiofilm', 'MDAD+aBiofilm_discard', 'aBiofilm_to_MDAD', 'Case_prediction',
                             'Ethnic_interaction'])
args = parser.parse_args()

def binary_cross_entropy(pred_output, labels):
    loss_fct = torch.nn.BCELoss()
    m = nn.Sigmoid()
    n = torch.squeeze(m(pred_output), 1)
    loss = loss_fct(n, labels)
    return n, loss

def cross_entropy_logits(linear_output, label, weights=None):
    class_output = F.log_softmax(linear_output, dim=1)
    n = F.softmax(linear_output, dim=1)[:, 1]
    max_class = class_output.max(1)
    y_hat = max_class[1]  # get the index of the max log-probability
    if weights is None:
        loss = nn.NLLLoss()(class_output, label.type_as(y_hat).view(label.size(0)))
    else:
        losses = nn.NLLLoss(reduction="none")(class_output, label.type_as(y_hat).view(label.size(0)))
        loss = torch.sum(weights * losses) / torch.sum(weights)
    return n, loss

cfg = get_cfg_defaults()
cfg.merge_from_file(args.cfg)
n_class = cfg["DECODER"]["BINARY"]
decoder = cfg["DECODER"]["NAME"]
params = {'batch_size': 1, 'shuffle': False, 'num_workers': cfg.SOLVER.NUM_WORKERS,
              'drop_last': True, 'collate_fn': graph_collate_func}
dataFolder = f'./datasets/{args.data}'
dataFolder = os.path.join(dataFolder, str(args.split))
result_path = os.path.join('./result/' + str(args.split), decoder)
test_path = os.path.join(dataFolder, "test_drugs.csv")
df_test = pd.read_csv(test_path)
test_dataset = DTIDataset(df_test.index.values, df_test)
test_generator = DataLoader(test_dataset, **params)
num_batches = len(test_generator)

MyModel = DrugBAN(**cfg).to(device)
MyModel.eval()
MyModel.load_state_dict(torch.load(result_path + '/best_model_epoch_19.pth')) #This place need to be confirm every time.

test_loss = 0
y_label_, y_pred_ = [], []
drugs_score = []
with torch.no_grad():
    for i, (v_d, v_p, a_m, d_i, labels) in enumerate(test_generator):
        a_m = torch.stack(a_m, 0)
        v_d, v_p, a_m, d_i, labels = v_d.to(device), v_p.to(device), a_m.to(device), d_i, labels.float().to(device)
        v_d, v_p, f, score = MyModel(v_d, v_p, a_m, d_i)
        drugs_score.append(v_d.tolist()[0])
        if n_class == 1:
            n, loss = binary_cross_entropy(score, labels)
        else:
            n, loss = cross_entropy_logits(score, labels)
        test_loss += loss.item()
        y_label_ = y_label_ + labels.to("cpu").tolist()
        y_pred_ = y_pred_ + n.to("cpu").tolist()

drug_name = ['Azithromycin', 'Daptomycin', 'Fosfomycin', 'Metronidazole', 'Vancomycin']
num_atoms = [52, 115, 8, 12, 101]

workbook = Workbook()
worksheet = workbook.active
for i in range(len(drugs_score)):
    if i == 0:
        worksheet.title = drug_name[i]
        for k in range(num_atoms[i]):
            worksheet.cell(row=1, column=1 + k).value = k
            for j in range(len(drugs_score[0][0])):
                worksheet.cell(row=2 + j, column=1 + k).value = drugs_score[i][k][j]
    else:
        worksheet1 = workbook.create_sheet(title=drug_name[i])
        for k in range(num_atoms[i]):
            worksheet1.cell(row=1, column=1 + k).value = k
            for j in range(len(drugs_score[0][0])):
                worksheet1.cell(row=2 + j, column=1 + k).value = drugs_score[i][k][j]
workbook.save('./Molecule_Graph_for_Drugs/Atom_vectors.xlsx')

drug_weight_matrix = np.loadtxt('./Molecule_Graph_for_Drugs/Drug_weight_matrix.txt')

SMILES_set = ['CCC1C(C(C(N(CC(CC(C(C(C(C(C(=O)O1)C)OC2CC(C(C(O2)C)O)(C)OC)C)OC3C(C(CC(O3)C)N(C)C)O)(C)O)C)C)C)O)(C)O',
              'CCCCCCCCCC(=O)NC(CC1=CNC2=CC=CC=C21)C(=O)NC(CC(=O)N)C(=O)NC(CC(=O)O)C(=O)NC3C(OC(=O)C(NC(=O)C(NC(=O)C(NC'
              '(=O)CNC(=O)C(NC(=O)C(NC(=O)C(NC(=O)C(NC(=O)CNC3=O)CCCN)CC(=O)O)C)CC(=O)O)CO)C(C)CC(=O)O)CC(=O)C4=CC=CC=C4N)C',
              'CC1C(O1)P(=O)(O)O', 'CC1=NC=C(N1CCO)[N+](=O)[O-]',
              'CC1C(C(CC(O1)OC2C(C(C(OC2OC3=C4C=C5C=C3OC6=C(C=C(C=C6)C(C(C(=O)NC(C(=O)NC5C(=O)NC7C8=CC(=C(C=C8)O)C9=C(C'
              '=C(C=C9O)O)C(NC(=O)C(C(C1=CC(=C(O4)C=C1)Cl)O)NC7=O)C(=O)O)CC(=O)N)NC(=O)C(CC(C)C)NC)O)Cl)CO)O)O)(C)N)O']
round_area_length = [40, 30, 100, 100, 30]
biases = [[7.5, -16], [-22, 0], [7, -5], [28, 38], [42, 30]]
stren_x = [1.25, 1.16, 1.15, 1.58, 1.51]
stren_y = [1.065, 1.12, 1.48, 1.34, 1.215]

for i in range(5):
    # SMILES string
    smiles = SMILES_set[i]

    # Convert SMILES to molecule object
    molecule = Chem.MolFromSmiles(smiles)

    if molecule.GetNumConformers() == 0:
        AllChem.Compute2DCoords(molecule)

    atom_scores = []
    drug_matrix = drug_weight_matrix @ np.array(drugs_score[i]).T


    min_x = float('inf')
    max_x = float('-inf')
    min_y = float('inf')
    max_y = float('-inf')
    for atom in molecule.GetAtoms():
        idx = atom.GetIdx()
        conf = molecule.GetConformer()
        pos = conf.GetAtomPosition(idx)
        print(f"Atom index: {idx}, Coordinates: ({pos.x}, {pos.y})")
        xx, yy = pos.x, pos.y
        min_x = min(min_x, xx)
        max_x = max(max_x, xx)
        min_y = min(min_y, yy)
        max_y = max(max_y, yy)
        atom_scores.append(np.linalg.norm(drug_matrix.T[idx], 1)/np.linalg.norm(np.array([drugs_score[i][idx]]), 'fro'))

    print(f"X coordinate range: {min_x} to {max_x}")
    print(f"Y coordinate range: {min_y} to {max_y}")
    lx = stren_x[i]*(max_x - min_x)/2
    ly = stren_y[i]*(max_y - min_y)/2

    img = Draw.MolToImage(molecule, size=(800, 600))

    pil_img = img.convert("RGBA")

    draw = ImageDraw.Draw(pil_img)

    atom_colors = {
        0: 'red',
        1: 'blue',
        2: 'green'
    }

    for atom_idx, color in atom_colors.items():
        atom = molecule.GetAtomWithIdx(atom_idx)

        conf = molecule.GetConformer()

        pos = conf.GetAtomPosition(atom_idx)

    gradient = Image.new('RGBA', (800, 600), (255, 255, 255, 0))
    draw_gradient = ImageDraw.Draw(gradient)

    end_color = (255, 165, 0, 0)

    r = round_area_length[i]//2
    for atom in molecule.GetAtoms():
        idx = atom.GetIdx()
        conf = molecule.GetConformer()
        pos = conf.GetAtomPosition(idx)
        depth = (atom_scores[idx]-min(atom_scores))/(max(atom_scores)-min(atom_scores))
        start_color = (255, 165, 0, 255*depth)
        stain_point = [int((400 + biases[i][0]) * (1 + pos.x / lx)), int((300 + biases[i][1]) * (1 - pos.y / ly))]
        for y in range(r, -r-1, -1):
            l = int((r*r-y*y)**0.5)
            for x in range(-l, l+1):
                ratio = (x*x+y*y)**0.5 / r
                color = tuple(int(b * ratio + a * (1 - ratio)) for a, b in zip(start_color, end_color))
                draw_gradient.point((stain_point[0]+x, stain_point[1]+y), fill=color)

    pil_img.paste(gradient, (0, 0), gradient)

    font = ImageFont.truetype("arial.ttf", 30)
    text = drug_name[i]
    text_width, text_height = font.getsize(text)
    draw.text((120, 10), text, fill="black", font=font)

    pil_img.show()
    pil_img.save('./Molecule_Graph_for_Drugs/Molecule_Graph_for_' + drug_name[i] + '.png')