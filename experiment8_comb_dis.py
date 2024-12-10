import csv
import random
import numpy as np
import openpyxl

st = ['SMILES', 'Microbe', 'Y']

# We selected some drugs that can find a few microbes with a few connected articles and confirm their connections
# one by one to build a test set that closer to the true meaning.
selected_d_1 = [1116, 444, 446, 574, 563, 557, 737, 465, 606, 599, 600, 902, 971]
selected_m_1_1 = [136, 141, 155, 153, 154, 65, 66, 67, 22, 21, 100, 101]
selected_m_1_2 = [68, 69, 94, 95, 96, 128, 129, 121, 123, 169, 171]
selected_d_2 = [766, 872, 1356, 1150, 445, 539, 872]
selected_m_2_1 = [28, 35, 36, 37, 38, 39, 40, 41, 42, 113]
selected_m_2_2 = [16, 17, 52, 73]
selected_d_3 = [421, 893, 926, 1192, 655, 1083, 1364]
selected_m_3_1 = [81, 83, 84, 85, 86, 87, 166]
selected_m_3_2 = [53, 88, 89, 90, 91, 92, 93]
print("Antibiotics-Gram positive bacteria:" + str(len(selected_d_1)*len(selected_m_1_1)))
print("Antibiotics-Gram negative bacteria:" + str(len(selected_d_1)*len(selected_m_1_2)))
print("Antifungal-Yeast:" + str(len(selected_d_2)*len(selected_m_2_1)))
print("Antifungal-Mold:" + str(len(selected_d_2)*len(selected_m_2_2)))
print("Antiviral-DNA_virus:" + str(len(selected_d_3)*len(selected_m_3_1)))
print("Antiviral-RNA_virus:" + str(len(selected_d_3)*len(selected_m_3_2)))

MDAD_drug_smiles = []
workbook = openpyxl.load_workbook('./source_dealed_data/MDAD_data/drug_microbe_matrix.xlsx')
sheet = workbook.active
dataa = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    MDAD_drug_smiles.append(row_data[1])
    dataa.append(row_data)
MDAD_microbe_name = dataa[0]

selected_area = []
selected_area_sp = []
for value in [[selected_d_1, selected_m_1_1, selected_m_1_2],
              [selected_d_2, selected_m_2_1, selected_m_2_2], [selected_d_3, selected_m_3_1, selected_m_3_2]]:
    for i in range(len(value[0])):
        for j in range(len(value[1])):
            selected_area.append([value[0][i], value[1][j]])
            selected_area_sp.append([dataa[value[0][i]][1], dataa[0][value[1][j] + 1]])
        for j in range(len(value[2])):
            selected_area.append([value[0][i], value[2][j]])
            selected_area_sp.append([dataa[value[0][i]][1], dataa[0][value[2][j] + 1]])

aBiofilm_drug_smiles = []
workbook = openpyxl.load_workbook('./source_dealed_data/aBiofilm_data/drug_microbe_matrix.xlsx')
sheet = workbook.active
data = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    aBiofilm_drug_smiles.append(row_data[1])
    data.append(row_data)
aBiofilm_microbe_name = data[0]

workbook = openpyxl.load_workbook('new_probably_connections(Fuzzy_set)/MDAD/connections.xlsx')
sheet = workbook.active
MD_probably_index = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    MD_probably_index.append([row_data[3], row_data[4]])
del MD_probably_index[0]

workbook = openpyxl.load_workbook('new_probably_connections(Fuzzy_set)/aBiofilm/connections.xlsx')
sheet = workbook.active
aB_probably_index = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    aB_probably_index.append([row_data[3], row_data[4]])
del aB_probably_index[0]

label_1 = []
label_0 = []

kn = np.loadtxt('./data_index/MDAD_index/known.txt')
md_kn = kn.tolist()
uk = np.loadtxt('./data_index/MDAD_index/unknown.txt')
kn = kn.tolist()
uk = uk.tolist()
for i in range(len(kn)):
    if kn[i] not in selected_area:
        label_1.append([dataa[int(kn[i][0])][1], dataa[0][int(kn[i][1]+1)], 1])
for i in range(len(uk)):
    if uk[i] not in selected_area:
        if uk[i] not in MD_probably_index:
            label_0.append([dataa[int(uk[i][0])][1], dataa[0][int(uk[i][1]+1)], 0])

kn = np.loadtxt('./data_index/aBiofilm_index/known.txt')
uk = np.loadtxt('./data_index/aBiofilm_index/unknown.txt')
kn = kn.tolist()
uk = uk.tolist()
aBiofilm_only_connect = []
aBiofilm_only_uncertain = []
for i in range(len(kn)):
    if data[int(kn[i][0])][1] not in MDAD_drug_smiles or data[0][int(kn[i][1] + 1)] not in MDAD_microbe_name:
        aBiofilm_only_connect.append([data[int(kn[i][0])][1], data[0][int(kn[i][1]+1)], 1])
for i in range(len(uk)):
    if data[int(uk[i][0])][1] not in MDAD_drug_smiles or data[0][int(uk[i][1] + 1)] not in MDAD_microbe_name:
        if uk[i] not in aB_probably_index:
            aBiofilm_only_uncertain.append([data[int(uk[i][0])][1],data[0][int(uk[i][1]+1)], 0])

label_1 = label_1 + aBiofilm_only_connect
label_0 = label_0 + aBiofilm_only_uncertain
random.shuffle(label_1)
random.shuffle(label_0)

kn_tr = label_1
uk_tr = label_0
val = []
test = []
for i in range(len(selected_area)):
    if selected_area[i] in md_kn:
        val.append([dataa[int(selected_area[i][0])][1], dataa[0][int(selected_area[i][1] + 1)], 1])
        test.append([dataa[int(selected_area[i][0])][1], dataa[0][int(selected_area[i][1] + 1)], 1])
    else:
        test.append([dataa[int(selected_area[i][0])][1], dataa[0][int(selected_area[i][1] + 1)], 0])
        if selected_area[i] not in MD_probably_index:
            val.append([dataa[int(selected_area[i][0])][1], dataa[0][int(selected_area[i][1] + 1)], 0])

train = kn_tr + uk_tr
random.shuffle(train)

with open('datasets/drug_microbe/Ethnic_interaction/val.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(st)
    for row in val:
        writer.writerow(row)
f.close()

with open('datasets/drug_microbe/Ethnic_interaction/train.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(st)
    for row in train:
        writer.writerow(row)
f.close()

with open('datasets/drug_microbe/Ethnic_interaction/test.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(st)
    for row in test:
        writer.writerow(row)
f.close()