import csv
import random
import numpy as np
import openpyxl

st = ['SMILES', 'Microbe', 'Y']

# We selected some drugs that can find a few microbes with a few connected articles and confirm their connections
# one by one to build a test set that closer to the true meaning.
selected_d = [88, 90, 91, 107, 109, 112, 113, 128, 149, 158, 160, 165, 202, 211, 221, 222, 224, 225, 229, 230, 231]
selected_m = [1, 2, 3, 4, 5, 6, 8, 10, 11, 12, 16, 18, 20, 22, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39,
              40, 41, 42, 43, 44, 49, 50, 51, 52, 56, 58, 60, 61, 63, 65, 66, 67, 68, 72, 73, 74, 76, 77, 78, 79,
              94, 95, 96, 97, 98, 99, 100, 103, 104, 105, 106, 108]

MDAD_drug_smiles = []
workbook = openpyxl.load_workbook('./source_dealed_data/MDAD_data/drug_microbe_matrix.xlsx')
sheet = workbook.active
dataa = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    MDAD_drug_smiles.append(row_data[1])
    dataa.append(row_data)
MDAD_microbe_name = dataa[0]

aBiofilm_drug_smiles = []
workbook = openpyxl.load_workbook('./source_dealed_data/aBiofilm_data/drug_microbe_matrix.xlsx')
sheet = workbook.active
data = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    aBiofilm_drug_smiles.append(row_data[1])
    data.append(row_data)
aBiofilm_microbe_name = data[0]

selected_drugs = []
selected_microbes = []
for i in range(len(selected_d)):
    selected_drugs.append(dataa[selected_d[i]][1])
for i in range(len(selected_m)):
    selected_microbes.append(dataa[0][selected_m[i]+1])

workbook = openpyxl.load_workbook('new_probably_connections(Fuzzy_set)/MDAD/connections.xlsx')
sheet = workbook.active
MD_probably_index = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    MD_probably_index.append([row_data[3], row_data[4]])
del MD_probably_index[0]

workbook = openpyxl.load_workbook('new_probably_connections(Fuzzy_set)/aBiofilm/connections.xlsx')
sheet = workbook.active
aB_probably_index = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    aB_probably_index.append([row_data[3], row_data[4]])
del aB_probably_index[0]

workbook = openpyxl.load_workbook('new_probably_connections(Fuzzy_set)/MDAD/connections0.xlsx')
sheet = workbook.active
MD_case_kn = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    if row_data[5] == 1:
        MD_case_kn.append([row_data[3], row_data[4]])
del MD_case_kn[0]

workbook = openpyxl.load_workbook('new_probably_connections(Fuzzy_set)/aBiofilm/connections0.xlsx')
sheet = workbook.active
aB_case_kn = []
for row in sheet.rows:
    row_data = []
    for cell in row:
        row_data.append(cell.value)
    if row_data[5] == 1:
        aB_case_kn.append([row_data[3], row_data[4]])
del aB_case_kn[0]

label_1 = []
label_0 = []

kn = np.loadtxt('./data_index/MDAD_index/known.txt')
md_kn = kn.tolist()
uk = np.loadtxt('./data_index/MDAD_index/unknown.txt')
kn = kn.tolist()
uk = uk.tolist()
for i in range(len(kn)):
    if kn[i][0] not in selected_d or kn[i][1] not in selected_m:
        label_1.append([dataa[int(kn[i][0])][1], dataa[0][int(kn[i][1]+1)], 1])
for i in range(len(uk)):
    if uk[i][0] not in selected_d or uk[i][1] not in selected_m:
        if uk[i] not in MD_case_kn:
            if uk[i] not in MD_probably_index:
                label_0.append([dataa[int(uk[i][0])][1], dataa[0][int(uk[i][1]+1)], 0])
        else:
            label_1.append([dataa[int(uk[i][0])][1], dataa[0][int(uk[i][1] + 1)], 1])

kn = np.loadtxt('./data_index/aBiofilm_index/known.txt')
uk = np.loadtxt('./data_index/aBiofilm_index/unknown.txt')
kn = kn.tolist()
uk = uk.tolist()
aBiofilm_only_connect = []
aBiofilm_only_uncertain = []
for i in range(len(kn)):
    if data[int(kn[i][0])][1] not in MDAD_drug_smiles or data[0][int(kn[i][1] + 1)] not in MDAD_microbe_name:
        aBiofilm_only_connect.append([data[int(kn[i][0])][1], data[0][int(kn[i][1]+1)], 1])
for i in range(len(uk)):
    if data[int(uk[i][0])][1] not in MDAD_drug_smiles or data[0][int(uk[i][1] + 1)] not in MDAD_microbe_name:
        if uk[i] not in aB_case_kn:
            if uk[i] not in aB_probably_index:
                aBiofilm_only_uncertain.append([data[int(uk[i][0])][1],data[0][int(uk[i][1]+1)], 0])
        else:
            aBiofilm_only_connect.append([data[int(uk[i][0])][1], data[0][int(uk[i][1] + 1)], 1])

label_1 = label_1 + aBiofilm_only_connect
label_0 = label_0 + aBiofilm_only_uncertain
random.shuffle(label_1)
random.shuffle(label_0)

kn_va = label_1[0:len(label_1)//8]
uk_va = label_0[0:len(label_0)//8]
kn_tr = label_1[len(label_1)//8:len(label_1)]
uk_tr = label_0[len(label_0)//8:len(label_0)]
kn_te = []
uk_te = []
for i in range(len(selected_d)):
    for j in range(len(selected_m)):
        if [selected_d[i], selected_m[j]] in MD_case_kn or [selected_d[i], selected_m[j]] in md_kn:
            kn_te.append([dataa[int(selected_d[i])][1], dataa[0][int(selected_m[j] + 1)], 1])
        else:
            uk_te.append([dataa[int(selected_d[i])][1], dataa[0][int(selected_m[j] + 1)], 0])
val = kn_va + uk_va
train = kn_tr + uk_tr
test = kn_te + uk_te
random.shuffle(val)
random.shuffle(train)

with open('datasets/drug_microbe/Case_prediction/val.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(st)
    for row in val:
        writer.writerow(row)
f.close()

with open('datasets/drug_microbe/Case_prediction/train.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(st)
    for row in train:
        writer.writerow(row)
f.close()

with open('datasets/drug_microbe/Case_prediction/test.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(st)
    for row in test:
        writer.writerow(row)
f.close()